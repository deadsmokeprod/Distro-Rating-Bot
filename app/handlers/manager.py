from __future__ import annotations

import asyncio
import logging
from datetime import date, datetime, timedelta
from math import ceil
import re
from pathlib import Path
import tempfile

from aiogram import F, Router
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import CallbackQuery, InlineKeyboardMarkup, Message
from aiogram.types import FSInputFile
from openpyxl import Workbook, load_workbook

from app.config import get_config
from app.db import sqlite
from app.handlers.start import is_admin, is_manager_or_admin, show_manager_menu
from app.handlers.filters import ActiveInlineMenuFilter, ManagerFilter, PrivateChatFilter
from app.keyboards.common import BACK_TEXT, build_inline_keyboard, support_contact_line, support_inline_keyboard
from app.keyboards.manager import (
    MANAGER_MENU_HELP,
    MANAGER_MENU_ORGS,
    MANAGER_MENU_REGISTER_ORG,
    MANAGER_MENU_EXPORT_RATINGS,
    MANAGER_MENU_BROADCAST,
    MANAGER_MENU_CHANGE_INN,
    MANAGER_MENU_MERGE_ORGS,
    MANAGER_MENU_GOALS_ADMIN,
    MANAGER_MENU_FIRE_ROP,
    MANAGER_MENU_RULES,
    MANAGER_MENU_SYNC,
    MANAGER_SYNC_CURRENT_MONTH,
    MANAGER_SYNC_CUSTOM_RANGE,
    MANAGER_BROADCAST_ALL,
    MANAGER_BROADCAST_BY_ORG,
    MANAGER_BROADCAST_MY_ORGS,
    MANAGER_BROADCAST_CONFIRM,
    ORG_ACTION_RESET_ROP_PASSWORD,
    ORG_ACTION_RESET_SELLER_PASSWORD,
    ORG_ACTION_STAFF,
    ORG_CREATE_BACK_TO_MENU,
    ORG_CREATE_CONFIRM,
    ORG_CREATE_OPEN_CARD,
    ORG_CREATE_OPEN_CARD_FULL,
    manager_back_menu,
    manager_main_menu,
    manager_broadcast_target_menu,
    manager_broadcast_confirm_menu,
    manager_goals_menu,
    manager_supertasks_menu,
    manager_avg_levels_menu,
    manager_sync_menu,
    org_create_confirm_menu,
    org_created_menu,
    org_exists_menu,
    GOALS_MENU_SUPERTASKS,
    GOALS_MENU_AVG_LEVELS,
    GOALS_MENU_DOWNLOAD_TEMPLATE,
    GOALS_MENU_UPLOAD_TEMPLATE,
    GOALS_MENU_AVG_CREATE,
)
from app.services.onec_client import OnecClientError
from app.services.turnover_sync import (
    current_month_range,
    moscow_today,
    send_sync_push_if_needed,
    sync_turnover,
)
from app.services.ratings import (
    month_str,
    moscow_today as moscow_today_ratings,
    current_month_rankings,
    get_all_time_for_user,
    get_monthly_snapshot_for_user,
    previous_month,
    recalc_all_time_ratings,
)
from app.services.ratings_export import build_ratings_excel
from app.services.leagues import compute_league
from app.services.challenges import get_current_challenge
from app.services.goals import sync_avg_levels_for_user
from app.utils.time import format_iso_human
from app.utils.security import generate_password, hash_password
from app.utils.validators import validate_inn, validate_org_name
from app.utils.inline_menu import clear_active_inline_menu, mark_inline_menu_active, send_single_inline_menu
from app.utils.rate_limit import is_rate_limited

logger = logging.getLogger(__name__)

router = Router()
router.message.filter(ManagerFilter())
router.callback_query.filter(ManagerFilter())
router.message.filter(PrivateChatFilter())
router.callback_query.filter(PrivateChatFilter())
router.callback_query.filter(ActiveInlineMenuFilter())

PAGE_SIZE = 10


def _can_access_org(actor_tg_user_id: int, org: object) -> bool:
    if is_admin(actor_tg_user_id):
        return True
    return int(org["created_by_manager_id"]) == actor_tg_user_id


def _manager_main_menu_for(actor_tg_user_id: int):
    return manager_main_menu(is_admin_view=is_admin(actor_tg_user_id))


def _person_label(full_name: str | None, tg_user_id: int) -> str:
    name = (full_name or "").strip()
    return f"{name} ({tg_user_id})" if name else f"ID {tg_user_id}"


def _row_full_name(row: dict | sqlite3.Row | None) -> str | None:
    if row is None:
        return None
    try:
        return row["full_name"]
    except Exception:
        return None


class OrgCreateStates(StatesGroup):
    inn = State()
    name = State()
    confirm = State()


class ManagerSyncStates(StatesGroup):
    choose_period = State()
    custom_range = State()


class ManagerExportStates(StatesGroup):
    period = State()


class ManagerBroadcastStates(StatesGroup):
    target = State()
    choose_org = State()
    message = State()
    confirm = State()


class AdminGoalsStates(StatesGroup):
    supertask_upload_wait_file = State()
    avg_level_wait_payload = State()


class ManagerInnChangeStates(StatesGroup):
    choose_org = State()
    old_inn = State()
    new_inn = State()
    confirm = State()


class AdminMergeStates(StatesGroup):
    choose_master = State()
    choose_joined = State()
    confirm_step1 = State()
    confirm_step2 = State()


async def _send_error(message: Message) -> None:
    await message.answer("Произошла ошибка, попробуйте позже.", reply_markup=manager_back_menu())


async def _send_secret_with_ttl(
    message: Message,
    text: str,
    ttl_sec: int = 180,
) -> None:
    sent = await message.answer(
        text + f"\n\nСообщение будет удалено через {ttl_sec} сек.",
        disable_web_page_preview=True,
    )

    async def _delete_later() -> None:
        await asyncio.sleep(ttl_sec)
        try:
            await message.bot.delete_message(chat_id=sent.chat.id, message_id=sent.message_id)
        except Exception:
            logger.debug("Failed to auto-delete secret message %s", sent.message_id)

    asyncio.create_task(_delete_later())


def _org_list_keyboard(
    orgs: list[dict], page: int, total_pages: int
) -> InlineKeyboardMarkup:
    buttons: list[tuple[str, str]] = []
    for org in orgs:
        text = f"{org['name']} — {org['inn']}"
        buttons.append((text, f"org_open:{org['id']}:{page}"))

    if page > 0:
        buttons.append(("◀️", f"org_page:{page - 1}"))
    if page < total_pages - 1:
        buttons.append(("▶️", f"org_page:{page + 1}"))
    buttons.append(("⬅️ Назад", "org_back_menu"))
    return build_inline_keyboard(buttons)


def _org_card_keyboard(org_id: int, back_page: int | None) -> InlineKeyboardMarkup:
    buttons = [
        (ORG_ACTION_STAFF, f"org_staff:{org_id}:0"),
        (ORG_ACTION_RESET_SELLER_PASSWORD, f"org_reset:{org_id}:seller"),
        (ORG_ACTION_RESET_ROP_PASSWORD, f"org_reset:{org_id}:rop"),
    ]
    if back_page is None:
        buttons.append(("⬅️ Назад", "org_back_menu"))
    else:
        buttons.append(("⬅️ Назад", f"org_page:{back_page}"))
    return build_inline_keyboard(buttons)


def _org_reset_confirm_keyboard(org_id: int, role: str) -> InlineKeyboardMarkup:
    role_label = "SELLER" if role == "seller" else "ROP"
    buttons = [
        (f"✅ Сбросить пароль {role_label}", f"org_reset_confirm:{org_id}:{role}"),
        ("⬅️ Назад", f"org_open:{org_id}:0"),
    ]
    return build_inline_keyboard(buttons)


def _org_staff_keyboard(
    org_id: int, page: int, total_pages: int, sellers: list
) -> InlineKeyboardMarkup:
    buttons: list[tuple[str, str]] = []
    for row in sellers:
        full_name = (row["full_name"] or "").strip()
        tg_user_id = int(row["tg_user_id"])
        label = f"{full_name} | {tg_user_id}" if full_name else f"ID {tg_user_id}"
        if len(label) > 64:
            label = label[:61] + "..."
        buttons.append((label, f"staff:{org_id}:{tg_user_id}:{page}"))
    if page > 0:
        buttons.append(("◀️", f"org_staff:{org_id}:{page - 1}"))
    if page < total_pages - 1:
        buttons.append(("▶️", f"org_staff:{org_id}:{page + 1}"))
    buttons.append(("⬅️ Назад", f"org_open:{org_id}:0"))
    return build_inline_keyboard(buttons)


def _inn_change_org_list_keyboard(
    orgs: list[dict], page: int, total_pages: int
) -> InlineKeyboardMarkup:
    buttons: list[tuple[str, str]] = []
    for org in orgs:
        buttons.append((f"{org['name']} — {org['inn']}", f"innchg_org_pick:{org['id']}:{page}"))
    if page > 0:
        buttons.append(("◀️", f"innchg_org_page:{page - 1}"))
    if page < total_pages - 1:
        buttons.append(("▶️", f"innchg_org_page:{page + 1}"))
    buttons.append(("⬅️ В меню", "org_back_menu"))
    return build_inline_keyboard(buttons)


async def _send_inn_change_org_list(
    message: Message, actor_tg_user_id: int, page: int, edit: bool = False
) -> None:
    config = get_config()
    if is_admin(actor_tg_user_id):
        total = await sqlite.count_orgs(config.db_path)
        orgs = await sqlite.list_orgs(config.db_path, PAGE_SIZE, max(0, page) * PAGE_SIZE)
    else:
        total = await sqlite.count_orgs_by_manager(config.db_path, actor_tg_user_id)
        orgs = await sqlite.list_orgs_by_manager(
            config.db_path, actor_tg_user_id, PAGE_SIZE, max(0, page) * PAGE_SIZE
        )
    if total <= 0:
        text = "Нет доступных организаций для смены ИНН."
        if edit:
            await message.edit_text(text, reply_markup=build_inline_keyboard([("⬅️ В меню", "org_back_menu")]))
            await mark_inline_menu_active(message, actor_tg_user_id)
        else:
            await message.answer(text, reply_markup=_manager_main_menu_for(actor_tg_user_id))
        return
    total_pages = max(1, ceil(total / PAGE_SIZE))
    page = max(0, min(page, total_pages - 1))
    if is_admin(actor_tg_user_id):
        orgs = await sqlite.list_orgs(config.db_path, PAGE_SIZE, page * PAGE_SIZE)
    else:
        orgs = await sqlite.list_orgs_by_manager(
            config.db_path, actor_tg_user_id, PAGE_SIZE, page * PAGE_SIZE
        )
    keyboard = _inn_change_org_list_keyboard([dict(o) for o in orgs], page, total_pages)
    text = "Выберите компанию для смены ИНН:"
    if edit:
        await message.edit_text(text, reply_markup=keyboard)
        await mark_inline_menu_active(message, actor_tg_user_id)
    else:
        await send_single_inline_menu(
            message,
            actor_tg_user_id=actor_tg_user_id,
            text=text,
            reply_markup=keyboard,
        )


def _broadcast_org_list_keyboard(
    orgs: list[dict], page: int, total_pages: int
) -> InlineKeyboardMarkup:
    buttons: list[tuple[str, str]] = []
    for org in orgs:
        buttons.append((f"{org['name']} — {org['inn']}", f"br_org_pick:{org['id']}:{page}"))
    if page > 0:
        buttons.append(("◀️", f"br_org_page:{page - 1}"))
    if page < total_pages - 1:
        buttons.append(("▶️", f"br_org_page:{page + 1}"))
    buttons.append(("⬅️ Назад", "br_org_back"))
    return build_inline_keyboard(buttons)


async def _send_broadcast_org_list(
    message: Message, actor_tg_user_id: int, page: int, edit: bool = False
) -> None:
    config = get_config()
    if is_admin(actor_tg_user_id):
        total = await sqlite.count_orgs(config.db_path)
    else:
        total = await sqlite.count_orgs_by_manager(config.db_path, actor_tg_user_id)
    if total <= 0:
        text = "Нет доступных компаний для адресной рассылки."
        if edit:
            await message.edit_text(
                text,
                reply_markup=build_inline_keyboard([("⬅️ В меню", "org_back_menu")]),
            )
            await mark_inline_menu_active(message, actor_tg_user_id)
        else:
            await message.answer(text, reply_markup=_manager_main_menu_for(actor_tg_user_id))
        return
    total_pages = max(1, ceil(total / PAGE_SIZE))
    page = max(0, min(page, total_pages - 1))
    if is_admin(actor_tg_user_id):
        orgs = await sqlite.list_orgs(config.db_path, PAGE_SIZE, page * PAGE_SIZE)
    else:
        orgs = await sqlite.list_orgs_by_manager(
            config.db_path, actor_tg_user_id, PAGE_SIZE, page * PAGE_SIZE
        )
    org_rows = [dict(o) for o in orgs]
    if edit:
        await message.edit_text(
            "Выберите компанию для рассылки:",
            reply_markup=_broadcast_org_list_keyboard(org_rows, page, total_pages),
        )
        await mark_inline_menu_active(message, actor_tg_user_id)
    else:
        await send_single_inline_menu(
            message,
            actor_tg_user_id=actor_tg_user_id,
            text="Выберите компанию для рассылки:",
            reply_markup=_broadcast_org_list_keyboard(org_rows, page, total_pages),
        )


def _merge_master_list_keyboard(orgs: list[dict], page: int, total_pages: int) -> InlineKeyboardMarkup:
    buttons: list[tuple[str, str]] = []
    for org in orgs:
        buttons.append((f"{org['name']} — {org['inn']}", f"merge_master_pick:{org['id']}:{page}"))
    if page > 0:
        buttons.append(("◀️", f"merge_master_page:{page - 1}"))
    if page < total_pages - 1:
        buttons.append(("▶️", f"merge_master_page:{page + 1}"))
    buttons.append(("⬅️ В меню", "org_back_menu"))
    return build_inline_keyboard(buttons)


def _merge_joined_list_keyboard(
    orgs: list[dict], selected_ids: set[int], page: int, total_pages: int
) -> InlineKeyboardMarkup:
    buttons: list[tuple[str, str]] = []
    for org in orgs:
        org_id = int(org["id"])
        mark = "✅" if org_id in selected_ids else "⬜"
        buttons.append((f"{mark} {org['name']} — {org['inn']}", f"merge_join_toggle:{org_id}:{page}"))
    if page > 0:
        buttons.append(("◀️", f"merge_join_page:{page - 1}"))
    if page < total_pages - 1:
        buttons.append(("▶️", f"merge_join_page:{page + 1}"))
    buttons.append(("✅ Продолжить", "merge_step1"))
    buttons.append(("🧹 Очистить выбор", "merge_clear"))
    buttons.append(("⬅️ В меню", "org_back_menu"))
    return build_inline_keyboard(buttons)


def _merge_confirm_step1_keyboard() -> InlineKeyboardMarkup:
    return build_inline_keyboard(
        [
            ("✅ Да, начать слияние", "merge_wait"),
            ("❌ Отмена", "merge_cancel"),
        ]
    )


def _merge_confirm_step2_keyboard() -> InlineKeyboardMarkup:
    return build_inline_keyboard(
        [
            ("✅ Подтвердить финально", "merge_execute"),
            ("❌ Отмена", "merge_cancel"),
        ]
    )


async def _send_merge_master_list(
    message: Message, actor_tg_user_id: int, page: int, edit: bool = False
) -> None:
    config = get_config()
    total = await sqlite.count_orgs(config.db_path)
    if total <= 1:
        text = "Для слияния нужно минимум 2 активные компании."
        kb = build_inline_keyboard([("⬅️ В меню", "org_back_menu")])
        if edit:
            await message.edit_text(text, reply_markup=kb)
            await mark_inline_menu_active(message, actor_tg_user_id)
        else:
            await send_single_inline_menu(
                message,
                actor_tg_user_id=actor_tg_user_id,
                text=text,
                reply_markup=kb,
            )
        return
    total_pages = max(1, ceil(total / PAGE_SIZE))
    page = max(0, min(page, total_pages - 1))
    orgs = [dict(r) for r in await sqlite.list_orgs(config.db_path, PAGE_SIZE, page * PAGE_SIZE)]
    kb = _merge_master_list_keyboard(orgs, page, total_pages)
    text = "Выберите мастер-компанию (куда вливаем):"
    if edit:
        await message.edit_text(text, reply_markup=kb)
        await mark_inline_menu_active(message, actor_tg_user_id)
    else:
        await send_single_inline_menu(
            message,
            actor_tg_user_id=actor_tg_user_id,
            text=text,
            reply_markup=kb,
        )


async def _send_merge_joined_list(
    message: Message,
    actor_tg_user_id: int,
    master_org_id: int,
    selected_ids: set[int],
    page: int,
    edit: bool = True,
) -> None:
    config = get_config()
    all_orgs = [dict(r) for r in await sqlite.list_orgs(config.db_path, 1000, 0)]
    candidates = [o for o in all_orgs if int(o["id"]) != master_org_id]
    total = len(candidates)
    if total <= 0:
        text = "Нет компаний для присоединения."
        kb = build_inline_keyboard([("⬅️ В меню", "org_back_menu")])
        if edit:
            await message.edit_text(text, reply_markup=kb)
            await mark_inline_menu_active(message, actor_tg_user_id)
        else:
            await send_single_inline_menu(
                message,
                actor_tg_user_id=actor_tg_user_id,
                text=text,
                reply_markup=kb,
            )
        return
    total_pages = max(1, ceil(total / PAGE_SIZE))
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    current = candidates[start:start + PAGE_SIZE]
    kb = _merge_joined_list_keyboard(current, selected_ids, page, total_pages)
    master = next((o for o in all_orgs if int(o["id"]) == master_org_id), None)
    master_title = f"{master['name']} — {master['inn']}" if master else str(master_org_id)
    text = (
        "Выберите одну или несколько компаний для присоединения:\n"
        f"Мастер: {master_title}\n"
        f"Выбрано: {len(selected_ids)}"
    )
    if edit:
        await message.edit_text(text, reply_markup=kb)
        await mark_inline_menu_active(message, actor_tg_user_id)
    else:
        await send_single_inline_menu(
            message,
            actor_tg_user_id=actor_tg_user_id,
            text=text,
            reply_markup=kb,
        )


def _parse_custom_range(text: str) -> tuple[date, date] | None:
    pattern = r"^\s*(\d{2})(\d{2})(\d{4})\s*по\s*(\d{2})(\d{2})(\d{4})\s*$"
    match = re.match(pattern, text)
    if not match:
        return None
    day1, month1, year1, day2, month2, year2 = match.groups()
    try:
        start = datetime(int(year1), int(month1), int(day1)).date()
        end = datetime(int(year2), int(month2), int(day2)).date()
    except ValueError:
        return None
    if start > end:
        return None
    if (end - start) > timedelta(days=60):
        return None
    return start, end


def _render_onec_error(exc: OnecClientError) -> str:
    text = f"❌ Ошибка 1С: {exc}"
    if getattr(exc, "hint", None):
        text += f"\nПодсказка: {exc.hint}"
    return text


def _parse_month_range(text: str) -> tuple[str, str] | None:
    pattern = r"^\s*с\s*(\d{2})\s*(\d{4})\s*по\s*(\d{2})\s*(\d{4})\s*$"
    match = re.match(pattern, text)
    if not match:
        return None
    m1, y1, m2, y2 = match.groups()
    try:
        start = date(int(y1), int(m1), 1)
        end = date(int(y2), int(m2), 1)
    except ValueError:
        return None
    if start > end:
        return None
    return month_str(start), month_str(end)


def _build_supertask_template() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "supertasks"
    ws.append(["region", "inn", "reward"])
    ws.append([77, "7707083893", 120])
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    wb.save(tmp.name)
    return Path(tmp.name)


def _parse_avg_level_payload(text: str) -> tuple[int, float, float, int] | None:
    # Формат: tg_user_id,target_liters,reward,days
    raw = [x.strip() for x in text.split(",")]
    if len(raw) != 4:
        return None
    try:
        tg_user_id = int(raw[0])
        target_liters = float(raw[1])
        reward = float(raw[2])
        days = int(raw[3])
    except ValueError:
        return None
    if tg_user_id <= 0 or target_liters <= 0 or reward <= 0 or days <= 0:
        return None
    return tg_user_id, target_liters, reward, days


def _broadcast_content_preview(content_type: str, text: str) -> str:
    if content_type == "text":
        preview = text.strip()
        if len(preview) > 120:
            preview = preview[:117] + "..."
        return f"Текст: {preview}" if preview else "Текстовое сообщение"
    labels = {
        "photo": "Фото",
        "video": "Видео",
        "voice": "Голосовое сообщение",
        "video_note": "Видео-кружок",
        "document": "Файл",
        "animation": "GIF/анимация",
        "sticker": "Стикер",
        "audio": "Аудио",
        "contact": "Контакт",
        "location": "Геолокация",
        "venue": "Точка/локация",
    }
    return labels.get(content_type, f"Контент ({content_type})")


def _is_service_message_type(content_type: str) -> bool:
    # Telegram service/system updates are not suitable for broadcast copy.
    service_types = {
        "new_chat_members",
        "left_chat_member",
        "new_chat_title",
        "new_chat_photo",
        "delete_chat_photo",
        "group_chat_created",
        "supergroup_chat_created",
        "channel_chat_created",
        "message_auto_delete_timer_changed",
        "migrate_to_chat_id",
        "migrate_from_chat_id",
        "pinned_message",
        "forum_topic_created",
        "forum_topic_edited",
        "forum_topic_closed",
        "forum_topic_reopened",
        "general_forum_topic_hidden",
        "general_forum_topic_unhidden",
        "video_chat_scheduled",
        "video_chat_started",
        "video_chat_ended",
        "video_chat_participants_invited",
        "write_access_allowed",
        "users_shared",
        "chat_shared",
        "connected_website",
        "passport_data",
        "proximity_alert_triggered",
        "web_app_data",
        "giveaway_created",
        "giveaway",
        "giveaway_winners",
        "giveaway_completed",
        "boost_added",
    }
    return content_type in service_types


@router.message(F.text == MANAGER_MENU_REGISTER_ORG)
async def manager_register_org(message: Message, state: FSMContext) -> None:
    if not is_manager_or_admin(message.from_user.id):
        return
    await state.clear()
    await state.set_state(OrgCreateStates.inn)
    await message.answer("Введите ИНН организации (10 или 12 цифр).", reply_markup=manager_back_menu())


@router.message(F.text == MANAGER_MENU_BROADCAST)
async def manager_broadcast_start(message: Message, state: FSMContext) -> None:
    await state.clear()
    await state.set_state(ManagerBroadcastStates.target)
    await message.answer(
        "Кому отправить сообщение?",
        reply_markup=manager_broadcast_target_menu(is_admin_view=is_admin(message.from_user.id)),
    )


@router.message(F.text == MANAGER_MENU_CHANGE_INN)
async def manager_change_inn_start(message: Message, state: FSMContext) -> None:
    if not is_manager_or_admin(message.from_user.id):
        return
    await state.clear()
    await state.set_state(ManagerInnChangeStates.choose_org)
    await _send_inn_change_org_list(message, actor_tg_user_id=message.from_user.id, page=0, edit=False)


@router.callback_query(F.data.startswith("innchg_org_page:"))
async def manager_change_inn_org_page(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    if not is_manager_or_admin(callback.from_user.id):
        return
    _, page_s = callback.data.split(":")
    await state.set_state(ManagerInnChangeStates.choose_org)
    await _send_inn_change_org_list(
        callback.message,
        actor_tg_user_id=callback.from_user.id,
        page=int(page_s),
        edit=True,
    )


@router.callback_query(F.data.startswith("innchg_org_pick:"))
async def manager_change_inn_org_pick(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    if not is_manager_or_admin(callback.from_user.id):
        return
    _, org_id_s, _page_s = callback.data.split(":")
    org_id = int(org_id_s)
    config = get_config()
    org = await sqlite.get_org_by_id(config.db_path, org_id)
    if not org or not _can_access_org(callback.from_user.id, org):
        await callback.message.edit_text(
            "Организация недоступна.",
            reply_markup=build_inline_keyboard([("⬅️ В меню", "org_back_menu")]),
        )
        return
    await state.set_state(ManagerInnChangeStates.old_inn)
    await state.update_data(org_id=org_id, org_name=str(org["name"]))
    inns = await sqlite.list_active_org_inns(config.db_path, org_id)
    inns_line = ", ".join(inns) if inns else "-"
    await callback.message.answer(
        "Введите текущий (старый) ИНН компании.\n"
        f"Активные ИНН: {inns_line}",
        reply_markup=manager_back_menu(),
    )


@router.message(ManagerInnChangeStates.old_inn, F.text == BACK_TEXT)
async def manager_change_inn_old_back(message: Message, state: FSMContext) -> None:
    await state.set_state(ManagerInnChangeStates.choose_org)
    await _send_inn_change_org_list(message, actor_tg_user_id=message.from_user.id, page=0, edit=False)


@router.message(ManagerInnChangeStates.old_inn)
async def manager_change_inn_old_input(message: Message, state: FSMContext) -> None:
    if not message.text:
        await message.answer("Введите ИНН или нажмите ⬅️ Назад.")
        return
    old_inn = message.text.strip()
    if not validate_inn(old_inn):
        await message.answer("ИНН должен содержать 10 или 12 цифр.", reply_markup=manager_back_menu())
        return
    data = await state.get_data()
    org_id = int(data.get("org_id", 0))
    config = get_config()
    if org_id <= 0 or not await sqlite.is_active_inn_for_org(config.db_path, org_id, old_inn):
        await message.answer("Такой активный ИНН не найден у выбранной компании.", reply_markup=manager_back_menu())
        return
    await state.update_data(old_inn=old_inn)
    await state.set_state(ManagerInnChangeStates.new_inn)
    await message.answer("Введите новый ИНН компании.", reply_markup=manager_back_menu())


@router.message(ManagerInnChangeStates.new_inn, F.text == BACK_TEXT)
async def manager_change_inn_new_back(message: Message, state: FSMContext) -> None:
    await state.set_state(ManagerInnChangeStates.old_inn)
    await message.answer("Введите текущий (старый) ИНН компании.", reply_markup=manager_back_menu())


@router.message(ManagerInnChangeStates.new_inn)
async def manager_change_inn_new_input(message: Message, state: FSMContext) -> None:
    if not message.text:
        await message.answer("Введите ИНН или нажмите ⬅️ Назад.")
        return
    new_inn = message.text.strip()
    if not validate_inn(new_inn):
        await message.answer("ИНН должен содержать 10 или 12 цифр.", reply_markup=manager_back_menu())
        return
    data = await state.get_data()
    old_inn = str(data.get("old_inn", "")).strip()
    org_id = int(data.get("org_id", 0))
    if org_id <= 0 or not old_inn:
        await state.clear()
        await show_manager_menu(message)
        return
    if new_inn == old_inn:
        await message.answer("Новый ИНН должен отличаться от текущего.", reply_markup=manager_back_menu())
        return
    config = get_config()
    existing_org = await sqlite.get_org_by_inn(config.db_path, new_inn)
    if existing_org and int(existing_org["id"]) != org_id:
        await message.answer("Новый ИНН уже используется другой активной компанией.", reply_markup=manager_back_menu())
        return
    await state.update_data(new_inn=new_inn)
    await state.set_state(ManagerInnChangeStates.confirm)
    await message.answer(
        "Подтвердите смену ИНН:\n"
        f"Старый ИНН: {old_inn}\n"
        f"Новый ИНН: {new_inn}",
        reply_markup=build_inline_keyboard(
            [
                ("✅ Подтвердить", "innchg_confirm_yes"),
                ("❌ Отмена", "innchg_confirm_no"),
            ]
        ),
    )


@router.callback_query(F.data == "innchg_confirm_no")
async def manager_change_inn_confirm_no(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    await state.clear()
    await callback.message.edit_text(
        "Смена ИНН отменена.",
        reply_markup=build_inline_keyboard([("⬅️ В меню", "org_back_menu")]),
    )


@router.callback_query(F.data == "innchg_confirm_yes")
async def manager_change_inn_confirm_yes(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    if not is_manager_or_admin(callback.from_user.id):
        return
    data = await state.get_data()
    org_id = int(data.get("org_id", 0))
    old_inn = str(data.get("old_inn", "")).strip()
    new_inn = str(data.get("new_inn", "")).strip()
    if org_id <= 0 or not old_inn or not new_inn:
        await state.clear()
        await callback.message.edit_text(
            "Недостаточно данных для смены ИНН.",
            reply_markup=build_inline_keyboard([("⬅️ В меню", "org_back_menu")]),
        )
        return
    config = get_config()
    org = await sqlite.get_org_by_id(config.db_path, org_id)
    if not org or not _can_access_org(callback.from_user.id, org):
        await state.clear()
        await callback.message.edit_text(
            "Организация недоступна.",
            reply_markup=build_inline_keyboard([("⬅️ В меню", "org_back_menu")]),
        )
        return
    changed = await sqlite.rotate_org_inn(config.db_path, org_id, old_inn, new_inn)
    if not changed:
        await state.clear()
        await callback.message.edit_text(
            "Не удалось выполнить смену ИНН. Проверьте данные и попробуйте снова.",
            reply_markup=build_inline_keyboard([("⬅️ В меню", "org_back_menu")]),
        )
        return
    await sqlite.log_audit(
        config.db_path,
        actor_tg_user_id=callback.from_user.id,
        actor_role="admin" if is_admin(callback.from_user.id) else "manager",
        action="ORG_INN_CHANGE",
        payload={"org_id": org_id, "old_inn": old_inn, "new_inn": new_inn},
    )
    await state.clear()
    await callback.message.edit_text(
        "ИНН успешно изменен.\n"
        f"Старый ИНН деактивирован: {old_inn}\n"
        f"Новый ИНН активирован: {new_inn}",
        reply_markup=build_inline_keyboard([("⬅️ В меню", "org_back_menu")]),
    )


@router.message(F.text == MANAGER_MENU_MERGE_ORGS)
async def manager_merge_start(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        await message.answer(
            "Слияние компаний доступно только ADMIN.",
            reply_markup=_manager_main_menu_for(message.from_user.id),
        )
        return
    await state.clear()
    await state.set_state(AdminMergeStates.choose_master)
    await _send_merge_master_list(message, actor_tg_user_id=message.from_user.id, page=0, edit=False)


@router.callback_query(F.data.startswith("merge_master_page:"))
async def manager_merge_master_page(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return
    _, page_s = callback.data.split(":")
    await state.set_state(AdminMergeStates.choose_master)
    await _send_merge_master_list(
        callback.message,
        actor_tg_user_id=callback.from_user.id,
        page=int(page_s),
        edit=True,
    )


@router.callback_query(F.data.startswith("merge_master_pick:"))
async def manager_merge_master_pick(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return
    _, master_org_id_s, _page_s = callback.data.split(":")
    master_org_id = int(master_org_id_s)
    config = get_config()
    master_org = await sqlite.get_org_by_id(config.db_path, master_org_id)
    if not master_org or int(master_org["is_active"]) != 1:
        await _send_merge_master_list(
            callback.message, actor_tg_user_id=callback.from_user.id, page=0, edit=True
        )
        return
    await state.set_state(AdminMergeStates.choose_joined)
    await state.update_data(merge_master_org_id=master_org_id, merge_joined_org_ids=[])
    await _send_merge_joined_list(
        callback.message,
        actor_tg_user_id=callback.from_user.id,
        master_org_id=master_org_id,
        selected_ids=set(),
        page=0,
        edit=True,
    )


@router.callback_query(F.data.startswith("merge_join_page:"))
async def manager_merge_join_page(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return
    _, page_s = callback.data.split(":")
    data = await state.get_data()
    master_org_id = int(data.get("merge_master_org_id", 0))
    selected = {int(x) for x in data.get("merge_joined_org_ids", [])}
    if master_org_id <= 0:
        await _send_merge_master_list(
            callback.message, actor_tg_user_id=callback.from_user.id, page=0, edit=True
        )
        return
    await _send_merge_joined_list(
        callback.message,
        actor_tg_user_id=callback.from_user.id,
        master_org_id=master_org_id,
        selected_ids=selected,
        page=int(page_s),
        edit=True,
    )


@router.callback_query(F.data.startswith("merge_join_toggle:"))
async def manager_merge_join_toggle(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return
    _, org_id_s, page_s = callback.data.split(":")
    org_id = int(org_id_s)
    page = int(page_s)
    data = await state.get_data()
    master_org_id = int(data.get("merge_master_org_id", 0))
    selected = {int(x) for x in data.get("merge_joined_org_ids", [])}
    if org_id in selected:
        selected.remove(org_id)
    else:
        selected.add(org_id)
    await state.update_data(merge_joined_org_ids=sorted(selected))
    await _send_merge_joined_list(
        callback.message,
        actor_tg_user_id=callback.from_user.id,
        master_org_id=master_org_id,
        selected_ids=selected,
        page=page,
        edit=True,
    )


@router.callback_query(F.data == "merge_clear")
async def manager_merge_clear(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return
    data = await state.get_data()
    master_org_id = int(data.get("merge_master_org_id", 0))
    await state.update_data(merge_joined_org_ids=[])
    await _send_merge_joined_list(
        callback.message,
        actor_tg_user_id=callback.from_user.id,
        master_org_id=master_org_id,
        selected_ids=set(),
        page=0,
        edit=True,
    )


@router.callback_query(F.data == "merge_step1")
async def manager_merge_step1(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return
    data = await state.get_data()
    master_org_id = int(data.get("merge_master_org_id", 0))
    joined = [int(x) for x in data.get("merge_joined_org_ids", [])]
    if master_org_id <= 0 or not joined:
        await callback.message.edit_text(
            "Нужно выбрать мастер-компанию и минимум одну присоединяемую.",
            reply_markup=build_inline_keyboard([("⬅️ Назад", "org_back_menu")]),
        )
        return
    config = get_config()
    all_orgs = {int(r["id"]): dict(r) for r in await sqlite.list_orgs(config.db_path, 1000, 0)}
    master = all_orgs.get(master_org_id)
    joined_names = [f"- {all_orgs[j]['name']} — {all_orgs[j]['inn']}" for j in joined if j in all_orgs]
    if not master:
        await _send_merge_master_list(
            callback.message, actor_tg_user_id=callback.from_user.id, page=0, edit=True
        )
        return
    await state.set_state(AdminMergeStates.confirm_step1)
    await callback.message.edit_text(
        "Подтвердите слияние компаний.\n"
        f"Мастер: {master['name']} — {master['inn']}\n"
        "Присоединяемые:\n"
        + ("\n".join(joined_names) if joined_names else "-"),
        reply_markup=_merge_confirm_step1_keyboard(),
    )


@router.callback_query(F.data == "merge_wait")
async def manager_merge_wait(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return
    await state.set_state(AdminMergeStates.confirm_step2)
    await callback.message.edit_text(
        "Финальное подтверждение доступно сразу.\nПроверьте выбранные компании и подтвердите действие.",
        reply_markup=_merge_confirm_step2_keyboard(),
    )


@router.callback_query(F.data == "merge_execute")
async def manager_merge_execute(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    if not is_admin(callback.from_user.id):
        return
    data = await state.get_data()
    master_org_id = int(data.get("merge_master_org_id", 0))
    joined = [int(x) for x in data.get("merge_joined_org_ids", [])]
    joined_key = ",".join(str(x) for x in sorted(joined))
    config = get_config()
    if is_rate_limited(
        f"merge_execute:{callback.from_user.id}",
        limit=config.merge_execute_limit,
        window_sec=config.merge_execute_window_sec,
    ):
        await callback.answer("Слишком много попыток слияния. Подождите немного.", show_alert=True)
        return
    if is_rate_limited(
        f"merge_execute_action:{callback.from_user.id}:{master_org_id}:{joined_key}",
        limit=1,
        window_sec=config.merge_execute_action_cooldown_sec,
    ):
        await callback.answer("Это слияние уже обрабатывается. Подождите немного.", show_alert=True)
        return
    if is_rate_limited(
        f"merge_execute_global:{callback.from_user.id}",
        limit=1,
        window_sec=config.merge_execute_global_cooldown_sec,
    ):
        await callback.answer(
            f"Новое слияние доступно через {config.merge_execute_global_cooldown_sec} сек.",
            show_alert=True,
        )
        return
    if master_org_id <= 0 or not joined:
        await state.clear()
        await callback.message.edit_text(
            "Данные слияния потеряны. Начните заново.",
            reply_markup=build_inline_keyboard([("⬅️ В меню", "org_back_menu")]),
        )
        return
    merged = await sqlite.merge_organizations(config.db_path, master_org_id, joined)
    if not merged:
        await state.clear()
        await callback.message.edit_text(
            "Не удалось выполнить слияние. Проверьте состояние компаний и повторите.",
            reply_markup=build_inline_keyboard([("⬅️ В меню", "org_back_menu")]),
        )
        return
    await sqlite.log_audit(
        config.db_path,
        actor_tg_user_id=callback.from_user.id,
        actor_role="admin",
        action="ORGS_MERGE",
        payload={"master_org_id": master_org_id, "joined_org_ids": joined},
    )
    await state.clear()
    await callback.message.edit_text(
        "Слияние выполнено необратимо.\n"
        "Присоединенные компании помечены merged/inactive,\n"
        "пользователи и ИНН объединены в мастер-группу.",
        reply_markup=build_inline_keyboard([("⬅️ В меню", "org_back_menu")]),
    )


@router.callback_query(F.data == "merge_cancel")
async def manager_merge_cancel(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    await state.clear()
    await callback.message.edit_text(
        "Слияние отменено.",
        reply_markup=build_inline_keyboard([("⬅️ В меню", "org_back_menu")]),
    )


@router.message(F.text == MANAGER_MENU_GOALS_ADMIN)
async def manager_goals_admin_open(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        await message.answer(
            "Раздел доступен только ADMIN.",
            reply_markup=_manager_main_menu_for(message.from_user.id),
        )
        return
    await state.clear()
    await message.answer("Админ-панель личных целей:", reply_markup=manager_goals_menu())


@router.message(F.text == GOALS_MENU_SUPERTASKS)
async def manager_goals_supertasks_open(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    await state.clear()
    await message.answer("Управление сверхзадачами:", reply_markup=manager_supertasks_menu())


@router.message(F.text == GOALS_MENU_DOWNLOAD_TEMPLATE)
async def manager_goals_download_template(message: Message) -> None:
    if not is_admin(message.from_user.id):
        return
    path: Path | None = None
    try:
        path = _build_supertask_template()
        await message.answer_document(
            FSInputFile(path, filename="supertasks_template.xlsx"),
            caption="Шаблон для загрузки сверхзадач.",
        )
    finally:
        if path is not None:
            try:
                path.unlink(missing_ok=True)
            except OSError:
                logger.warning("Failed to remove temporary template file: %s", path)


@router.message(F.text == GOALS_MENU_UPLOAD_TEMPLATE)
async def manager_goals_upload_template_start(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    await state.set_state(AdminGoalsStates.supertask_upload_wait_file)
    await message.answer("Пришлите Excel-файл со столбцами: region, inn, reward.", reply_markup=manager_back_menu())


@router.message(AdminGoalsStates.supertask_upload_wait_file, F.text == BACK_TEXT)
async def manager_goals_upload_back(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer("Управление сверхзадачами:", reply_markup=manager_supertasks_menu())


@router.message(AdminGoalsStates.supertask_upload_wait_file, F.document)
async def manager_goals_upload_template_file(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    config = get_config()
    doc = message.document
    if not doc:
        await message.answer("Файл не найден. Попробуйте снова.")
        return
    tmp_path: Path | None = None
    try:
        file = await message.bot.get_file(doc.file_id)
        stream = await message.bot.download_file(file.file_path)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.write(stream.read())
        tmp.close()
        tmp_path = Path(tmp.name)
        wb = load_workbook(tmp_path, data_only=True)
        ws = wb.active
        created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            region_raw = row[0] if len(row) > 0 else None
            inn_raw = row[1] if len(row) > 1 else None
            reward_raw = row[2] if len(row) > 2 else None
            if region_raw is None or inn_raw is None or reward_raw is None:
                continue
            try:
                region = int(float(region_raw))
                inn = str(inn_raw).strip()
                reward = float(reward_raw)
            except (TypeError, ValueError):
                continue
            if region <= 0 or reward <= 0 or not validate_inn(inn):
                continue
            await sqlite.create_supertask(
                config.db_path,
                region=region,
                target_inn=inn,
                reward=reward,
                created_by_tg_user_id=message.from_user.id,
            )
            created += 1
        await sqlite.log_audit(
            config.db_path,
            actor_tg_user_id=message.from_user.id,
            actor_role="admin",
            action="SUPERTASKS_UPLOAD",
            payload={"created": created},
        )
        if created > 0 and config.supertask_push_new_enabled:
            recipients = await sqlite.list_sellers_and_rops_active(config.db_path)
            for recipient in recipients:
                tg_user_id = int(recipient["tg_user_id"])
                try:
                    await message.bot.send_message(
                        tg_user_id,
                        f"Добавлены новые сверхзадачи: {created} шт. Откройте раздел «🎯 Личные цели».",
                    )
                except Exception:
                    logger.exception("Failed supertask new push to %s", tg_user_id)
        await state.clear()
        await message.answer(f"Загрузка завершена. Создано сверхзадач: {created}", reply_markup=manager_supertasks_menu())
    except Exception:
        logger.exception("Failed to upload supertasks excel")
        await message.answer("Не удалось обработать файл. Проверьте формат и попробуйте снова.")
    finally:
        if tmp_path is not None:
            try:
                tmp_path.unlink(missing_ok=True)
            except OSError:
                logger.warning("Failed to remove uploaded temporary file: %s", tmp_path)


@router.message(F.text == GOALS_MENU_AVG_LEVELS)
async def manager_goals_avg_open(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    await state.clear()
    await message.answer("Управление уровнями среднемесячного:", reply_markup=manager_avg_levels_menu())


@router.message(F.text == GOALS_MENU_AVG_CREATE)
async def manager_goals_avg_create_start(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    await state.set_state(AdminGoalsStates.avg_level_wait_payload)
    await message.answer(
        "Введите параметры через запятую:\n"
        "tg_user_id,target_liters,reward,days\n"
        "Пример: 123456789,250,80,30",
        reply_markup=manager_back_menu(),
    )


@router.message(AdminGoalsStates.avg_level_wait_payload, F.text == BACK_TEXT)
async def manager_goals_avg_create_back(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer("Управление уровнями среднемесячного:", reply_markup=manager_avg_levels_menu())


@router.message(AdminGoalsStates.avg_level_wait_payload)
async def manager_goals_avg_create_submit(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    if not message.text:
        await message.answer("Введите данные или нажмите ⬅️ Назад.")
        return
    parsed = _parse_avg_level_payload(message.text)
    if not parsed:
        await message.answer(
            "Неверный формат. Ожидается:\n"
            "tg_user_id,target_liters,reward,days",
            reply_markup=manager_back_menu(),
        )
        return
    tg_user_id, target_liters, reward, days = parsed
    config = get_config()
    user = await sqlite.get_user_by_tg_id(config.db_path, tg_user_id)
    if not user or str(user["status"]) != "active" or str(user["role"]) not in {"seller", "rop"}:
        await message.answer("Пользователь не найден или неактивен.")
        return
    active_levels = await sqlite.count_active_levels_for_user(config.db_path, tg_user_id)
    if active_levels >= config.max_avg_levels:
        await message.answer(f"Достигнут лимит активных уровней ({config.max_avg_levels}).")
        return
    starts_at = datetime.utcnow().isoformat()
    ends_at = (datetime.utcnow() + timedelta(days=days)).isoformat()
    avg_level_id = await sqlite.create_avg_level(
        config.db_path,
        tg_user_id=tg_user_id,
        target_liters=target_liters,
        reward=reward,
        starts_at=starts_at,
        ends_at=ends_at,
        created_by_tg_user_id=message.from_user.id,
    )
    await sync_avg_levels_for_user(config, tg_user_id)
    await sqlite.log_audit(
        config.db_path,
        actor_tg_user_id=message.from_user.id,
        actor_role="admin",
        action="AVG_LEVEL_CREATE",
        payload={
            "avg_level_id": avg_level_id,
            "tg_user_id": tg_user_id,
            "target_liters": target_liters,
            "reward": reward,
            "days": days,
        },
    )
    await state.clear()
    user_label = _person_label(_row_full_name(user), tg_user_id)
    await message.answer(
        f"Уровень создан (#{avg_level_id}) для {user_label}.",
        reply_markup=manager_avg_levels_menu(),
    )


@router.message(ManagerBroadcastStates.target, F.text == BACK_TEXT)
async def manager_broadcast_back(message: Message, state: FSMContext) -> None:
    await state.clear()
    await show_manager_menu(message)


@router.message(
    ManagerBroadcastStates.target,
    F.text.in_({MANAGER_BROADCAST_ALL, MANAGER_BROADCAST_MY_ORGS, MANAGER_BROADCAST_BY_ORG}),
)
async def manager_broadcast_target(message: Message, state: FSMContext) -> None:
    text = message.text or ""
    if text == MANAGER_BROADCAST_ALL:
        if not is_admin(message.from_user.id):
            await message.answer(
                "Режим «Всем продавцам» доступен только ADMIN.",
                reply_markup=manager_broadcast_target_menu(is_admin_view=False),
            )
            return
        await state.update_data(target="all", target_org_id=None, target_org_name=None)
        await state.set_state(ManagerBroadcastStates.message)
        await message.answer(
            "Отправьте сообщение для рассылки.\n"
            "Поддерживаются: текст, фото, видео, голосовые, кружки, файлы и др.",
            reply_markup=manager_back_menu(),
        )
        return
    if text == MANAGER_BROADCAST_MY_ORGS:
        await state.update_data(target="my_orgs", target_org_id=None, target_org_name=None)
        await state.set_state(ManagerBroadcastStates.message)
        await message.answer(
            "Отправьте сообщение для рассылки.\n"
            "Поддерживаются: текст, фото, видео, голосовые, кружки, файлы и др.",
            reply_markup=manager_back_menu(),
        )
        return
    await state.update_data(target="org", target_org_id=None, target_org_name=None)
    await state.set_state(ManagerBroadcastStates.choose_org)
    await _send_broadcast_org_list(
        message,
        actor_tg_user_id=message.from_user.id,
        page=0,
        edit=False,
    )


@router.callback_query(ManagerBroadcastStates.choose_org, F.data == "br_org_back")
async def manager_broadcast_org_back(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    await state.set_state(ManagerBroadcastStates.target)
    await callback.message.answer(
        "Кому отправить сообщение?",
        reply_markup=manager_broadcast_target_menu(is_admin_view=is_admin(callback.from_user.id)),
    )


@router.message(ManagerBroadcastStates.choose_org, F.text == BACK_TEXT)
async def manager_broadcast_org_back_text(message: Message, state: FSMContext) -> None:
    await state.set_state(ManagerBroadcastStates.target)
    await message.answer(
        "Кому отправить сообщение?",
        reply_markup=manager_broadcast_target_menu(is_admin_view=is_admin(message.from_user.id)),
    )


@router.callback_query(ManagerBroadcastStates.choose_org, F.data.startswith("br_org_page:"))
async def manager_broadcast_org_page(callback: CallbackQuery) -> None:
    await callback.answer()
    _, page_s = callback.data.split(":")
    await _send_broadcast_org_list(
        callback.message,
        actor_tg_user_id=callback.from_user.id,
        page=int(page_s),
        edit=True,
    )


@router.callback_query(ManagerBroadcastStates.choose_org, F.data.startswith("br_org_pick:"))
async def manager_broadcast_org_pick(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    _, org_id_s, _page_s = callback.data.split(":")
    org_id = int(org_id_s)
    config = get_config()
    org = await sqlite.get_org_by_id(config.db_path, org_id)
    if not org or not _can_access_org(callback.from_user.id, org):
        await callback.answer("Организация недоступна.", show_alert=True)
        return
    await state.update_data(
        target="org",
        target_org_id=org_id,
        target_org_name=str(org["name"]),
    )
    await state.set_state(ManagerBroadcastStates.message)
    await callback.message.answer(
        f"Выбрана компания: {org['name']} ({org['inn']}).\n"
        "Отправьте сообщение для рассылки (текст или любой поддерживаемый медиа-тип).",
        reply_markup=manager_back_menu(),
    )


@router.message(ManagerBroadcastStates.message, F.text == BACK_TEXT)
async def manager_broadcast_message_back(message: Message, state: FSMContext) -> None:
    data = await state.get_data()
    if data.get("target") == "org":
        await state.set_state(ManagerBroadcastStates.choose_org)
        await _send_broadcast_org_list(
            message,
            actor_tg_user_id=message.from_user.id,
            page=0,
            edit=False,
        )
        return
    await state.set_state(ManagerBroadcastStates.target)
    await message.answer(
        "Кому отправить сообщение?",
        reply_markup=manager_broadcast_target_menu(is_admin_view=is_admin(message.from_user.id)),
    )


@router.message(ManagerBroadcastStates.message)
async def manager_broadcast_message(message: Message, state: FSMContext) -> None:
    content_type = str(message.content_type or "")
    if _is_service_message_type(content_type):
        await message.answer(
            "Сервисные сообщения Telegram нельзя использовать в рассылке.\n"
            "Отправьте обычный контент: текст, фото, видео, кружок, файл, контакт, гео и т.д."
        )
        return
    text_payload = (message.text or message.caption or "").strip()
    await state.update_data(
        source_chat_id=message.chat.id,
        source_message_id=message.message_id,
        content_type=content_type,
        text=text_payload,
    )
    await state.set_state(ManagerBroadcastStates.confirm)
    data = await state.get_data()
    target = str(data.get("target", "all"))
    content_preview = _broadcast_content_preview(content_type, text_payload)
    if target == "org":
        org_name = str(data.get("target_org_name") or "не выбрана")
        prompt = f"Отправить это сообщение ({content_preview}) в компанию «{org_name}»?"
    elif target == "my_orgs":
        prompt = f"Отправить это сообщение ({content_preview}) продавцам ваших компаний?"
    else:
        prompt = f"Отправить это сообщение ({content_preview}) всем продавцам?"
    await message.answer(prompt, reply_markup=manager_broadcast_confirm_menu())


@router.message(ManagerBroadcastStates.confirm, F.text == BACK_TEXT)
async def manager_broadcast_confirm_back(message: Message, state: FSMContext) -> None:
    await state.set_state(ManagerBroadcastStates.message)
    await message.answer(
        "Отправьте сообщение для рассылки (текст или медиа).",
        reply_markup=manager_back_menu(),
    )


@router.message(ManagerBroadcastStates.confirm, F.text == MANAGER_BROADCAST_CONFIRM)
async def manager_broadcast_send(message: Message, state: FSMContext) -> None:
    data = await state.get_data()
    target = data.get("target")
    source_chat_id = data.get("source_chat_id")
    source_message_id = data.get("source_message_id")
    content_type = str(data.get("content_type") or "")
    if source_chat_id is None or source_message_id is None:
        await message.answer(
            "Сообщение для рассылки не выбрано.",
            reply_markup=_manager_main_menu_for(message.from_user.id),
        )
        await state.clear()
        return
    config = get_config()
    if target == "all":
        if not is_admin(message.from_user.id):
            await message.answer(
                "Режим «Всем продавцам» доступен только ADMIN.",
                reply_markup=manager_broadcast_target_menu(is_admin_view=False),
            )
            await state.set_state(ManagerBroadcastStates.target)
            return
        recipients = await sqlite.list_all_seller_ids(config.db_path)
        target_meta = {"target": "all"}
    elif target == "org":
        org_id = int(data.get("target_org_id") or 0)
        org = await sqlite.get_org_by_id(config.db_path, org_id) if org_id > 0 else None
        if not org or not _can_access_org(message.from_user.id, org):
            await message.answer(
                "Компания для рассылки недоступна. Выберите получателей заново.",
                reply_markup=manager_broadcast_target_menu(is_admin_view=is_admin(message.from_user.id)),
            )
            await state.set_state(ManagerBroadcastStates.target)
            return
        recipients = await sqlite.list_seller_ids_by_org(config.db_path, org_id)
        target_meta = {"target": "org", "org_id": org_id, "org_name": str(org["name"])}
    else:
        recipients = await sqlite.list_seller_ids_by_manager(config.db_path, message.from_user.id)
        target_meta = {"target": "my_orgs"}
    sent = 0
    for tg_user_id in recipients:
        try:
            await message.bot.copy_message(
                chat_id=tg_user_id,
                from_chat_id=int(source_chat_id),
                message_id=int(source_message_id),
            )
            sent += 1
        except Exception:
            logger.exception("Failed to send broadcast to %s", tg_user_id)
            continue
    await sqlite.log_audit(
        config.db_path,
        actor_tg_user_id=message.from_user.id,
        actor_role="admin" if is_admin(message.from_user.id) else "manager",
        action="MANAGER_BROADCAST",
        payload={**target_meta, "content_type": content_type, "recipients": len(recipients), "sent": sent},
    )
    await state.clear()
    await message.answer(
        f"Рассылка завершена. Отправлено: {sent}",
        reply_markup=_manager_main_menu_for(message.from_user.id),
    )


@router.message(F.text == MANAGER_MENU_EXPORT_RATINGS)
async def manager_export_ratings_start(message: Message, state: FSMContext) -> None:
    await state.clear()
    await state.set_state(ManagerExportStates.period)
    await message.answer(
        'Введите период в формате: "с ММ ГГГГ по ММ ГГГГ".\n'
        "Например: с 01 2026 по 03 2026",
        reply_markup=manager_back_menu(),
    )


@router.message(ManagerExportStates.period, F.text == BACK_TEXT)
async def manager_export_ratings_back(message: Message, state: FSMContext) -> None:
    await state.clear()
    await show_manager_menu(message)


@router.message(ManagerExportStates.period)
async def manager_export_ratings_run(message: Message, state: FSMContext) -> None:
    if not message.text:
        await message.answer("Введите период или нажмите ⬅️ Назад.")
        return
    parsed = _parse_month_range(message.text)
    if not parsed:
        await message.answer(
            'Неверный формат. Ожидаю: "с ММ ГГГГ по ММ ГГГГ".',
            reply_markup=manager_back_menu(),
        )
        return
    start_month, end_month = parsed
    await state.clear()
    config = get_config()
    current_month_label = f"{moscow_today_ratings().month:02d} {moscow_today_ratings().year}"
    path: Path | None = None
    try:
        path = await build_ratings_excel(
            config.db_path, start_month, end_month, current_month_label
        )
        filename = f"ratings_{start_month}_to_{end_month}.xlsx"
        await message.answer_document(
            FSInputFile(path, filename=filename),
            caption="Выгрузка рейтингов",
        )
    except Exception:
        logger.exception("Failed to export ratings")
        await message.answer(
            "Не удалось сформировать выгрузку.",
            reply_markup=_manager_main_menu_for(message.from_user.id),
        )
    finally:
        if path is not None:
            try:
                path.unlink(missing_ok=True)
            except OSError:
                logger.warning("Failed to remove temporary ratings export file: %s", path)


@router.message(F.text == MANAGER_MENU_SYNC)
async def manager_sync_start(message: Message, state: FSMContext) -> None:
    await state.clear()
    await state.set_state(ManagerSyncStates.choose_period)
    await message.answer("Выберите период для обновления базы.", reply_markup=manager_sync_menu())


@router.message(ManagerSyncStates.choose_period, F.text == BACK_TEXT)
async def manager_sync_back(message: Message, state: FSMContext) -> None:
    await state.clear()
    await show_manager_menu(message)


@router.message(ManagerSyncStates.choose_period, F.text == MANAGER_SYNC_CURRENT_MONTH)
async def manager_sync_current_month(message: Message, state: FSMContext) -> None:
    config = get_config()
    await state.clear()
    start, end = current_month_range(moscow_today())
    operation_type = config.onec_operation_type
    try:
        await message.answer("Запрос к 1С отправлен, ожидайте…")
        sync_result = await sync_turnover(config, start, end, operation_type=operation_type)
        push_sent = await send_sync_push_if_needed(message.bot, config, sync_result)
        await sqlite.log_audit(
            config.db_path,
            actor_tg_user_id=message.from_user.id,
            actor_role="manager",
            action="SYNC_TURNOVER",
            payload={
                "mode": "current_month",
                "operationType": operation_type,
                "start": start.isoformat(),
                "end": end.isoformat(),
                "fetched": sync_result.fetched_count,
                "upserted": sync_result.upserted_count,
                "inserted_count": sync_result.inserted_count,
                "affected_company_group_ids": sync_result.affected_company_group_ids,
                "sync_push_sent": push_sent,
            },
        )
        await message.answer(
            "✅ Обновление завершено.\n"
            f"Период: {start.isoformat()} — {end.isoformat()}\n"
            f"Получено строк из 1С: {sync_result.fetched_count}\n"
            f"Записано/обновлено в базу: {sync_result.upserted_count}\n"
            f"Новых строк: {sync_result.inserted_count}\n"
            f"Push-уведомлений отправлено: {push_sent}",
            reply_markup=_manager_main_menu_for(message.from_user.id),
        )
    except OnecClientError as exc:
        logger.warning(
            "Turnover sync failed (current month): status=%s code=%s actor=%s",
            getattr(exc, "status_code", None),
            getattr(exc, "code", "ONEC_ERROR"),
            message.from_user.id,
        )
        try:
            await sqlite.log_audit(
                config.db_path,
                actor_tg_user_id=message.from_user.id,
                actor_role="manager",
                action="SYNC_TURNOVER_ERROR",
                payload={
                    "mode": "current_month",
                    "operationType": operation_type,
                    "start": start.isoformat(),
                    "end": end.isoformat(),
                    "status_code": getattr(exc, "status_code", None),
                    "error_code": getattr(exc, "code", "ONEC_ERROR"),
                    "error": str(exc),
                },
            )
        except Exception:
            logger.exception("Failed to write sync error audit (current month)")
        await message.answer(
            _render_onec_error(exc),
            reply_markup=_manager_main_menu_for(message.from_user.id),
        )
    except Exception as exc:
        logger.exception("Failed to sync turnover (current month)")
        await message.answer(
            f"❌ Ошибка обновления базы: {exc}",
            reply_markup=_manager_main_menu_for(message.from_user.id),
        )


@router.message(ManagerSyncStates.choose_period, F.text == MANAGER_SYNC_CUSTOM_RANGE)
async def manager_sync_custom_range_start(message: Message, state: FSMContext) -> None:
    await state.set_state(ManagerSyncStates.custom_range)
    await message.answer(
        "Введите период в формате: ДДММГГГГ по ДДММГГГГ.\n"
        "Например: 01012026 по 31012026",
        reply_markup=manager_back_menu(),
    )


@router.message(ManagerSyncStates.custom_range, F.text == BACK_TEXT)
async def manager_sync_custom_back(message: Message, state: FSMContext) -> None:
    await state.set_state(ManagerSyncStates.choose_period)
    await message.answer("Выберите период для обновления базы.", reply_markup=manager_sync_menu())


@router.message(ManagerSyncStates.custom_range)
async def manager_sync_custom_range(message: Message, state: FSMContext) -> None:
    if not message.text:
        await message.answer("Введите период или нажмите ⬅️ Назад.")
        return
    parsed = _parse_custom_range(message.text)
    if not parsed:
        await message.answer(
            "Неверный формат или период больше 60 дней.\n"
            "Ожидаю: ДДММГГГГ по ДДММГГГГ.",
            reply_markup=manager_back_menu(),
        )
        return
    start, end = parsed
    config = get_config()
    operation_type = config.onec_operation_type
    await state.clear()
    try:
        await message.answer("Запрос к 1С отправлен, ожидайте…")
        sync_result = await sync_turnover(config, start, end, operation_type=operation_type)
        push_sent = await send_sync_push_if_needed(message.bot, config, sync_result)
        await sqlite.log_audit(
            config.db_path,
            actor_tg_user_id=message.from_user.id,
            actor_role="manager",
            action="SYNC_TURNOVER",
            payload={
                "mode": "custom_range",
                "operationType": operation_type,
                "start": start.isoformat(),
                "end": end.isoformat(),
                "fetched": sync_result.fetched_count,
                "upserted": sync_result.upserted_count,
                "inserted_count": sync_result.inserted_count,
                "affected_company_group_ids": sync_result.affected_company_group_ids,
                "sync_push_sent": push_sent,
            },
        )
        await message.answer(
            "✅ Обновление завершено.\n"
            f"Период: {start.isoformat()} — {end.isoformat()}\n"
            f"Получено строк из 1С: {sync_result.fetched_count}\n"
            f"Записано/обновлено в базу: {sync_result.upserted_count}\n"
            f"Новых строк: {sync_result.inserted_count}\n"
            f"Push-уведомлений отправлено: {push_sent}",
            reply_markup=_manager_main_menu_for(message.from_user.id),
        )
    except OnecClientError as exc:
        logger.warning(
            "Turnover sync failed (custom range): status=%s code=%s actor=%s",
            getattr(exc, "status_code", None),
            getattr(exc, "code", "ONEC_ERROR"),
            message.from_user.id,
        )
        try:
            await sqlite.log_audit(
                config.db_path,
                actor_tg_user_id=message.from_user.id,
                actor_role="manager",
                action="SYNC_TURNOVER_ERROR",
                payload={
                    "mode": "custom_range",
                    "operationType": operation_type,
                    "start": start.isoformat(),
                    "end": end.isoformat(),
                    "status_code": getattr(exc, "status_code", None),
                    "error_code": getattr(exc, "code", "ONEC_ERROR"),
                    "error": str(exc),
                },
            )
        except Exception:
            logger.exception("Failed to write sync error audit (custom range)")
        await message.answer(
            _render_onec_error(exc),
            reply_markup=_manager_main_menu_for(message.from_user.id),
        )
    except Exception as exc:
        logger.exception("Failed to sync turnover (custom range)")
        await message.answer(
            f"❌ Ошибка обновления базы: {exc}",
            reply_markup=_manager_main_menu_for(message.from_user.id),
        )


@router.message(OrgCreateStates.inn, F.text == BACK_TEXT)
async def manager_org_inn_back(message: Message, state: FSMContext) -> None:
    await state.clear()
    await show_manager_menu(message)


@router.message(OrgCreateStates.inn)
async def manager_org_inn_input(message: Message, state: FSMContext) -> None:
    if not message.text:
        await message.answer("Пожалуйста, введите ИНН или нажмите ⬅️ Назад.")
        return
    inn = message.text.strip()
    if not validate_inn(inn):
        await message.answer("ИНН должен содержать 10 или 12 цифр", reply_markup=manager_back_menu())
        return

    config = get_config()
    try:
        existing = await sqlite.get_org_by_inn(config.db_path, inn)
        if existing:
            if int(existing["created_by_manager_id"]) == message.from_user.id:
                await state.clear()
                await state.update_data(existing_org_id=int(existing["id"]))
                await message.answer(
                    "Организация уже зарегистрирована вами.", reply_markup=org_exists_menu()
                )
                return
            await state.clear()
            await message.answer("Организация уже зарегистрирована.", reply_markup=manager_back_menu())
            return

        await state.update_data(inn=inn)
        await state.set_state(OrgCreateStates.name)
        await message.answer("Введите наименование организации.", reply_markup=manager_back_menu())
    except Exception:
        logger.exception("Failed to handle org inn input")
        await _send_error(message)


@router.message(OrgCreateStates.name, F.text == BACK_TEXT)
async def manager_org_name_back(message: Message, state: FSMContext) -> None:
    await state.set_state(OrgCreateStates.inn)
    await message.answer("Введите ИНН организации (10 или 12 цифр).", reply_markup=manager_back_menu())


@router.message(OrgCreateStates.name)
async def manager_org_name_input(message: Message, state: FSMContext) -> None:
    if not message.text:
        await message.answer("Пожалуйста, введите наименование или нажмите ⬅️ Назад.")
        return
    name = message.text.strip()
    if not validate_org_name(name):
        await message.answer(
            "Наименование должно быть от 2 до 200 символов.", reply_markup=manager_back_menu()
        )
        return
    data = await state.get_data()
    inn = data.get("inn")
    await state.update_data(name=name)
    await state.set_state(OrgCreateStates.confirm)
    await message.answer(
        f"Проверьте данные:\nИНН: {inn}\nНаименование: {name}\nСоздать организацию?",
        reply_markup=org_create_confirm_menu(),
    )


@router.message(OrgCreateStates.confirm, F.text == BACK_TEXT)
async def manager_org_confirm_back(message: Message, state: FSMContext) -> None:
    await state.set_state(OrgCreateStates.name)
    await message.answer("Введите наименование организации.", reply_markup=manager_back_menu())


@router.message(OrgCreateStates.confirm, F.text == ORG_CREATE_CONFIRM)
async def manager_org_confirm_create(message: Message, state: FSMContext) -> None:
    data = await state.get_data()
    inn = data.get("inn")
    name = data.get("name")
    if not inn or not name:
        await state.clear()
        await show_manager_menu(message)
        return

    config = get_config()
    try:
        seller_password_plain = generate_password()
        rop_password_plain = generate_password()
        org_id = await sqlite.create_org(
            config.db_path,
            inn=inn,
            name=name,
            seller_password_hash=hash_password(seller_password_plain),
            rop_password_hash=hash_password(rop_password_plain),
            created_by_manager_id=message.from_user.id,
        )
        await sqlite.log_audit(
            config.db_path,
            actor_tg_user_id=message.from_user.id,
            actor_role="manager",
            action="ORG_CREATE",
            payload={"org_id": org_id, "inn": inn},
        )
        await state.clear()
        await state.update_data(existing_org_id=org_id)
        await message.answer(
            "Организация создана.\n"
            f"ИНН: {inn}\n"
            f"Наименование: {name}\n"
            "Пароли отправлены отдельным сообщением с авто-удалением.",
            reply_markup=org_created_menu(),
        )
        await _send_secret_with_ttl(
            message,
            "Секретные пароли организации:\n"
            f"SELLER: {seller_password_plain}\n"
            f"ROP: {rop_password_plain}",
        )
    except Exception:
        logger.exception("Failed to create org")
        await _send_error(message)


@router.message(OrgCreateStates.confirm)
async def manager_org_confirm_fallback(message: Message) -> None:
    await message.answer("Пожалуйста, нажмите ✅ Создать или ⬅️ Назад.")


@router.message(F.text == ORG_CREATE_OPEN_CARD)
@router.message(F.text == ORG_CREATE_OPEN_CARD_FULL)
async def manager_open_card_from_message(message: Message, state: FSMContext) -> None:
    if not is_manager_or_admin(message.from_user.id):
        return
    data = await state.get_data()
    org_id = data.get("existing_org_id")
    if not org_id:
        await show_manager_menu(message)
        return
    await _send_org_card(message, message.from_user.id, org_id, back_page=None)


@router.message(F.text == ORG_CREATE_BACK_TO_MENU)
async def manager_back_to_menu(message: Message, state: FSMContext) -> None:
    await state.clear()
    await show_manager_menu(message)


@router.message(F.text == MANAGER_MENU_ORGS)
async def manager_org_list(message: Message) -> None:
    if not is_manager_or_admin(message.from_user.id):
        return
    await _send_org_list(message, actor_tg_user_id=message.from_user.id, page=0)


@router.callback_query(F.data.startswith("org_page"))
async def manager_org_list_page(callback: CallbackQuery) -> None:
    if not is_manager_or_admin(callback.from_user.id):
        await callback.answer()
        return
    if callback.message is None:
        await callback.answer()
        return
    parts = callback.data.split(":")
    if len(parts) != 2:
        await callback.answer()
        return
    try:
        page = int(parts[1])
    except ValueError:
        await callback.answer()
        return
    await _send_org_list(
        callback.message, actor_tg_user_id=callback.from_user.id, page=page, edit=True
    )
    await callback.answer()


@router.callback_query(F.data == "org_back_menu")
async def manager_org_back_menu(callback: CallbackQuery) -> None:
    if not is_manager_or_admin(callback.from_user.id):
        await callback.answer()
        return
    role_name = "Администратор" if is_admin(callback.from_user.id) else "Менеджер"
    await clear_active_inline_menu(callback.message, callback.from_user.id)
    await callback.message.answer(
        f"Вы вошли как {role_name}.",
        reply_markup=_manager_main_menu_for(callback.from_user.id),
    )
    await callback.answer()


async def _send_org_list(
    message: Message, actor_tg_user_id: int, page: int, edit: bool = False
) -> None:
    config = get_config()
    if is_admin(actor_tg_user_id):
        total = await sqlite.count_orgs(config.db_path)
    else:
        total = await sqlite.count_orgs_by_manager(config.db_path, actor_tg_user_id)
    total_pages = max(1, ceil(total / PAGE_SIZE))
    page = max(0, min(page, total_pages - 1))
    if is_admin(actor_tg_user_id):
        orgs = await sqlite.list_orgs(config.db_path, PAGE_SIZE, page * PAGE_SIZE)
    else:
        orgs = await sqlite.list_orgs_by_manager(
            config.db_path, actor_tg_user_id, PAGE_SIZE, page * PAGE_SIZE
        )
    keyboard = _org_list_keyboard(orgs, page, total_pages)
    text = "Ваши организации:" if total else "У вас пока нет организаций."
    if edit:
        await message.edit_text(text, reply_markup=keyboard)
        await mark_inline_menu_active(message, actor_tg_user_id)
    else:
        await send_single_inline_menu(
            message,
            actor_tg_user_id=actor_tg_user_id,
            text=text,
            reply_markup=keyboard,
        )


async def _send_org_card(
    message: Message, user_id: int, org_id: int, back_page: int | None
) -> None:
    config = get_config()
    org = await sqlite.get_org_by_id(config.db_path, org_id)
    if not org or not _can_access_org(user_id, org):
        await message.answer("Организация не найдена.", reply_markup=_manager_main_menu_for(user_id))
        return
    await sqlite.log_audit(
        config.db_path,
        actor_tg_user_id=user_id,
        actor_role="manager",
        action="VIEW_ORG",
        payload={"org_id": org_id},
    )
    seller_count = await sqlite.count_sellers_by_org(config.db_path, org_id)
    rop_count = await sqlite.count_active_rops_by_org(config.db_path, org_id)
    text = (
        "Организация:\n"
        f"ИНН: {org['inn']}\n"
        f"Наименование: {org['name']}\n"
        f"Активных SELLER: {seller_count}\n"
        f"Активных ROP: {rop_count}"
    )
    keyboard = _org_card_keyboard(org_id, back_page)
    await send_single_inline_menu(
        message,
        actor_tg_user_id=user_id,
        text=text,
        reply_markup=keyboard,
    )


@router.callback_query(F.data.startswith("org_open"))
async def manager_org_open(callback: CallbackQuery) -> None:
    if not is_manager_or_admin(callback.from_user.id):
        await callback.answer()
        return
    if callback.message is None:
        await callback.answer()
        return
    parts = callback.data.split(":")
    if len(parts) != 3:
        await callback.answer()
        return
    try:
        _, org_id_s, page_s = parts
        org_id = int(org_id_s)
        page = int(page_s)
    except ValueError:
        await callback.answer()
        return
    await _send_org_card(callback.message, callback.from_user.id, org_id, back_page=page)
    await callback.answer()


def _escape_html(text: str) -> str:
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


@router.callback_query(F.data.startswith("staff:"))
async def manager_staff_profile(callback: CallbackQuery) -> None:
    if not is_manager_or_admin(callback.from_user.id):
        await callback.answer()
        return
    parts = callback.data.split(":")
    if len(parts) != 4:
        await callback.answer()
        return
    _, org_id_s, tg_user_id_s, staff_page_s = parts
    org_id = int(org_id_s)
    target_tg_user_id = int(tg_user_id_s)
    staff_page = int(staff_page_s)
    config = get_config()
    org = await sqlite.get_org_by_id(config.db_path, org_id)
    if not org or not _can_access_org(callback.from_user.id, org):
        await callback.message.answer(
            "Организация не найдена.",
            reply_markup=_manager_main_menu_for(callback.from_user.id),
        )
        await callback.answer()
        return
    user = await sqlite.get_user_by_tg_id(config.db_path, target_tg_user_id)
    if not user or int(user["org_id"]) != org_id:
        await callback.answer("Сотрудник не найден.")
        return
    await recalc_all_time_ratings(config.db_path)
    all_time = await get_all_time_for_user(config.db_path, target_tg_user_id) or {
        "total_volume": 0,
        "global_rank": 0,
        "company_rank": 0,
    }
    today = moscow_today_ratings()
    prev_month = previous_month(today)
    prev_snapshot = await get_monthly_snapshot_for_user(
        config.db_path, prev_month, target_tg_user_id
    ) or {"total_volume": 0, "global_rank": 0, "company_rank": 0}
    league = compute_league(
        await current_month_rankings(config.db_path), target_tg_user_id
    )
    challenge = await get_current_challenge(config, target_tg_user_id)
    challenge_line = ""
    if challenge:
        if challenge.completed:
            challenge_line = "Испытание месяца пройдено ✅\n"
        else:
            challenge_line = (
                f"Испытание месяца: {challenge.progress_volume:g}/{challenge.target_volume:g} л\n"
            )
    league_line = f"Лига: {league.name}"
    if league.to_next_volume is not None:
        league_line += f", до повышения {league.to_next_volume:g} л"
    registered_at = format_iso_human(user["registered_at"])
    has_req = await sqlite.has_requisites(config.db_path, target_tg_user_id)
    requisites_line = "Реквизиты указаны: Да" if has_req else "Реквизиты указаны: Нет"
    user_label = _person_label(_row_full_name(user), target_tg_user_id)
    profile_text = (
        "Профиль сотрудника:\n"
        f"Сотрудник: {_escape_html(user_label)}\n"
        f"Дата регистрации: {registered_at}\n"
        f"{requisites_line}\n\n"
        + challenge_line
        + league_line
        + "\n\n"
        "Рейтинг за всё время: "
        f"{all_time['total_volume']} (в прошлом месяце было {prev_snapshot['total_volume']})\n"
        "Место в мировом рейтинге: "
        f"{all_time['global_rank']} (в прошлом месяце было {prev_snapshot['global_rank']})\n"
        "Место в рейтинге компании: "
        f"{all_time['company_rank']} (в прошлом месяце было {prev_snapshot['company_rank']})"
    )
    history = await sqlite.get_requisites_history(config.db_path, target_tg_user_id)
    if history:
        profile_text += "\n\n——— История реквизитов ———\n"
        for i, row in enumerate(history):
            dt = format_iso_human(row["created_at"])
            content = _escape_html(str(row["content"]))
            if i == 0:
                profile_text += f"\n<b>{dt}</b>\n<b>{content}</b>\n"
            else:
                profile_text += f"\n{dt}\n{content}\n"
    else:
        profile_text += "\n\n——— История реквизитов ———\nНет записей."
    back_kb = build_inline_keyboard([
        ("⬅️ Назад к списку", f"org_staff:{org_id}:{staff_page}"),
    ])
    await callback.message.edit_text(
        profile_text,
        reply_markup=back_kb,
        parse_mode="HTML",
    )
    await callback.answer()


@router.callback_query(F.data.startswith("org_staff"))
async def manager_org_staff(callback: CallbackQuery) -> None:
    if not is_manager_or_admin(callback.from_user.id):
        await callback.answer()
        return
    _, org_id, page = callback.data.split(":")
    org_id = int(org_id)
    page = int(page)
    config = get_config()
    org = await sqlite.get_org_by_id(config.db_path, org_id)
    if not org or not _can_access_org(callback.from_user.id, org):
        await callback.message.answer(
            "Организация не найдена.",
            reply_markup=_manager_main_menu_for(callback.from_user.id),
        )
        await callback.answer()
        return
    total = await sqlite.count_sellers_by_org(config.db_path, org_id)
    total_pages = max(1, ceil(total / PAGE_SIZE))
    page = max(0, min(page, total_pages - 1))
    sellers = await sqlite.list_sellers_by_org(
        config.db_path, org_id, PAGE_SIZE, page * PAGE_SIZE
    )
    if sellers:
        text = "Сотрудники (нажмите на сотрудника для просмотра профиля и истории реквизитов):"
    else:
        text = "Сотрудники не зарегистрированы."
    keyboard = _org_staff_keyboard(org_id, page, total_pages, sellers)
    await callback.message.edit_text(text, reply_markup=keyboard)
    await callback.answer()


@router.callback_query(F.data.startswith("org_reset:"))
async def manager_org_reset(callback: CallbackQuery) -> None:
    if not is_manager_or_admin(callback.from_user.id):
        await callback.answer()
        return
    parts = callback.data.split(":")
    if len(parts) != 3:
        await callback.answer("Некорректные данные.")
        return
    _, org_id, role = parts
    if role not in {"seller", "rop"}:
        await callback.answer("Неизвестный тип пароля.")
        return
    await callback.message.edit_text(
        f"Сбросить пароль {role.upper()}? Старый перестанет работать.",
        reply_markup=_org_reset_confirm_keyboard(int(org_id), role),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("org_reset_confirm:"))
async def manager_org_reset_confirm(callback: CallbackQuery) -> None:
    if not is_manager_or_admin(callback.from_user.id):
        await callback.answer()
        return
    parts = callback.data.split(":")
    if len(parts) != 3:
        await callback.answer("Некорректные данные.")
        return
    _, org_id, role = parts
    if role not in {"seller", "rop"}:
        await callback.answer("Неизвестный тип пароля.")
        return
    org_id_int = int(org_id)
    config = get_config()
    org = await sqlite.get_org_by_id(config.db_path, org_id_int)
    if not org or not _can_access_org(callback.from_user.id, org):
        await callback.message.answer(
            "Организация не найдена.",
            reply_markup=_manager_main_menu_for(callback.from_user.id),
        )
        await callback.answer()
        return
    password_plain = generate_password()
    password_hash = hash_password(password_plain)
    await sqlite.update_org_password(config.db_path, org_id_int, role, password_hash)
    await sqlite.log_audit(
        config.db_path,
        actor_tg_user_id=callback.from_user.id,
        actor_role="manager",
        action="ORG_PASSWORD_RESET",
        payload={"org_id": org_id_int, "role": role},
    )
    await callback.message.edit_text(
        f"Новый пароль {role.upper()} отправлен отдельным сообщением с авто-удалением.",
        reply_markup=build_inline_keyboard([(BACK_TEXT, f"org_open:{org_id_int}:0")]),
    )
    await _send_secret_with_ttl(
        callback.message,
        f"Новый пароль {role.upper()}: {password_plain}",
    )
    await callback.answer()


def _fire_rop_orgs_keyboard(orgs: list[dict]) -> InlineKeyboardMarkup:
    buttons: list[tuple[str, str]] = []
    for org in orgs:
        title = f"{org['name']} — {org['inn']}"
        buttons.append((title, f"fire_rop_org:{org['id']}"))
    buttons.append(("⬅️ Назад", "org_back_menu"))
    return build_inline_keyboard(buttons)


def _rop_action_list_keyboard(org_id: int, rops: list[dict], action: str) -> InlineKeyboardMarkup:
    buttons: list[tuple[str, str]] = []
    for row in rops:
        name = (row["full_name"] or "").strip() or f"ID {row['tg_user_id']}"
        buttons.append((f"{name} | {row['tg_user_id']}", f"rop_{action}:{org_id}:{row['tg_user_id']}"))
    buttons.append(("⬅️ Назад", f"fire_rop_org:{org_id}"))
    return build_inline_keyboard(buttons)


@router.message(F.text == MANAGER_MENU_FIRE_ROP)
async def manager_fire_rop_menu(message: Message) -> None:
    if not is_manager_or_admin(message.from_user.id):
        return
    config = get_config()
    if is_admin(message.from_user.id):
        orgs = await sqlite.list_orgs(config.db_path, 100, 0)
    else:
        orgs = await sqlite.list_orgs_by_manager(config.db_path, message.from_user.id, 100, 0)
    org_list = [dict(r) for r in orgs]
    if not org_list:
        await message.answer(
            "Нет доступных организаций.",
            reply_markup=_manager_main_menu_for(message.from_user.id),
        )
        return
    await message.answer("Выберите организацию для управления РОП:", reply_markup=_fire_rop_orgs_keyboard(org_list))


@router.callback_query(F.data.startswith("fire_rop_org:"))
async def manager_fire_rop_org(callback: CallbackQuery) -> None:
    await callback.answer()
    if not is_manager_or_admin(callback.from_user.id):
        return
    _, org_id_s = callback.data.split(":")
    org_id = int(org_id_s)
    config = get_config()
    org = await sqlite.get_org_by_id(config.db_path, org_id)
    if not org or not _can_access_org(callback.from_user.id, org):
        await callback.message.answer(
            "Организация не найдена.",
            reply_markup=_manager_main_menu_for(callback.from_user.id),
        )
        return
    active_rops = [dict(r) for r in await sqlite.list_active_rops_by_org(config.db_path, org_id)]
    fired_rops = [dict(r) for r in await sqlite.list_fired_rops_by_org(config.db_path, org_id)]
    buttons = [
        (f"Уволить действующего РОП ({len(active_rops)})", f"fire_rop_list:{org_id}:active"),
        (f"Уволенные РОП ({len(fired_rops)})", f"fire_rop_list:{org_id}:fired"),
        ("⬅️ Назад", "org_back_menu"),
    ]
    await callback.message.edit_text("Выберите действие:", reply_markup=build_inline_keyboard(buttons))


@router.callback_query(F.data.startswith("fire_rop_list:"))
async def manager_fire_rop_list(callback: CallbackQuery) -> None:
    await callback.answer()
    if not is_manager_or_admin(callback.from_user.id):
        return
    _, org_id_s, mode = callback.data.split(":")
    org_id = int(org_id_s)
    config = get_config()
    if mode == "active":
        rows = [dict(r) for r in await sqlite.list_active_rops_by_org(config.db_path, org_id)]
        if not rows:
            await callback.message.edit_text(
                "В этой организации нет активных РОП.",
                reply_markup=build_inline_keyboard([("⬅️ Назад", f"fire_rop_org:{org_id}")]),
            )
            return
        await callback.message.edit_text(
            "Выберите РОП для увольнения:",
            reply_markup=_rop_action_list_keyboard(org_id, rows, "fire"),
        )
        return
    rows = [dict(r) for r in await sqlite.list_fired_rops_by_org(config.db_path, org_id)]
    if not rows:
        await callback.message.edit_text(
            "Нет уволенных РОП для восстановления.",
            reply_markup=build_inline_keyboard([("⬅️ Назад", f"fire_rop_org:{org_id}")]),
        )
        return
    await callback.message.edit_text(
        "Выберите РОП для восстановления:",
        reply_markup=_rop_action_list_keyboard(org_id, rows, "restore"),
    )


@router.callback_query(F.data.startswith("rop_fire:"))
async def manager_fire_rop_confirm(callback: CallbackQuery) -> None:
    await callback.answer()
    if not is_manager_or_admin(callback.from_user.id):
        return
    _, org_id_s, tg_user_id_s = callback.data.split(":")
    org_id = int(org_id_s)
    tg_user_id = int(tg_user_id_s)
    config = get_config()
    org = await sqlite.get_org_by_id(config.db_path, org_id)
    if not org or not _can_access_org(callback.from_user.id, org):
        await callback.message.edit_text(
            "Организация недоступна.",
            reply_markup=build_inline_keyboard([("⬅️ В меню", "org_back_menu")]),
        )
        return
    target_user = await sqlite.get_user_by_tg_id(config.db_path, tg_user_id)
    if not target_user or int(target_user["org_id"]) != org_id:
        await callback.message.edit_text(
            "РОП не найден в выбранной организации.",
            reply_markup=build_inline_keyboard([("⬅️ Назад", f"fire_rop_org:{org_id}")]),
        )
        return
    changed = await sqlite.fire_user(
        config.db_path,
        tg_user_id=tg_user_id,
        expected_role="rop",
        fired_by_tg_user_id=callback.from_user.id,
    )
    if not changed:
        await callback.message.edit_text(
            "Не удалось уволить РОП (возможно, статус уже изменился).",
            reply_markup=build_inline_keyboard([("⬅️ Назад", f"fire_rop_org:{org_id}")]),
        )
        return
    await sqlite.log_audit(
        config.db_path,
        actor_tg_user_id=callback.from_user.id,
        actor_role="manager",
        action="FIRE_ROP",
        payload={"org_id": org_id, "tg_user_id": tg_user_id},
    )
    await callback.message.edit_text(
        "РОП уволен (soft).",
        reply_markup=build_inline_keyboard([("⬅️ Назад", f"fire_rop_org:{org_id}")]),
    )


@router.callback_query(F.data.startswith("rop_restore:"))
async def manager_restore_rop_confirm(callback: CallbackQuery) -> None:
    await callback.answer()
    if not is_manager_or_admin(callback.from_user.id):
        return
    _, org_id_s, tg_user_id_s = callback.data.split(":")
    org_id = int(org_id_s)
    tg_user_id = int(tg_user_id_s)
    config = get_config()
    org = await sqlite.get_org_by_id(config.db_path, org_id)
    if not org or not _can_access_org(callback.from_user.id, org):
        await callback.message.edit_text(
            "Организация недоступна.",
            reply_markup=build_inline_keyboard([("⬅️ В меню", "org_back_menu")]),
        )
        return
    target_user = await sqlite.get_user_by_tg_id(config.db_path, tg_user_id)
    if not target_user or int(target_user["org_id"]) != org_id:
        await callback.message.edit_text(
            "РОП не найден в выбранной организации.",
            reply_markup=build_inline_keyboard([("⬅️ Назад", f"fire_rop_org:{org_id}")]),
        )
        return
    if await sqlite.has_active_registration_in_other_org(config.db_path, tg_user_id, org_id):
        await callback.message.edit_text(
            "Восстановление невозможно: у пользователя активная регистрация в другой компании.",
            reply_markup=build_inline_keyboard([("⬅️ Назад", f"fire_rop_org:{org_id}")]),
        )
        return
    changed = await sqlite.restore_user(
        config.db_path,
        tg_user_id=tg_user_id,
        expected_role="rop",
    )
    if not changed:
        await callback.message.edit_text(
            "Не удалось восстановить РОП.",
            reply_markup=build_inline_keyboard([("⬅️ Назад", f"fire_rop_org:{org_id}")]),
        )
        return
    await sqlite.log_audit(
        config.db_path,
        actor_tg_user_id=callback.from_user.id,
        actor_role="manager",
        action="RESTORE_ROP",
        payload={"org_id": org_id, "tg_user_id": tg_user_id},
    )
    await callback.message.edit_text(
        "РОП восстановлен.",
        reply_markup=build_inline_keyboard([("⬅️ Назад", f"fire_rop_org:{org_id}")]),
    )


@router.message(F.text == MANAGER_MENU_HELP)
async def manager_help(message: Message) -> None:
    if not is_manager_or_admin(message.from_user.id):
        return
    config = get_config()
    await message.answer(
        "Бот помогает регистрировать организации и продавцов.\n"
        "Если возникли вопросы — напишите в техподдержку."
        + support_contact_line(config.support_username),
        reply_markup=support_inline_keyboard(config.support_user_id, config.support_username),
    )


@router.message(F.text == MANAGER_MENU_RULES)
async def manager_rules(message: Message) -> None:
    if not is_manager_or_admin(message.from_user.id):
        return
    config = get_config()
    rules_path = Path(config.rules_file_path)
    if not rules_path.exists() or not rules_path.is_file():
        await message.answer(
            "Файл с правилами пока недоступен. Обратитесь в техподдержку."
            + support_contact_line(config.support_username),
            reply_markup=support_inline_keyboard(config.support_user_id, config.support_username),
        )
        return
    await message.answer_document(
        FSInputFile(rules_path),
        caption="Правила и рекомендации.",
    )


@router.message(F.text == BACK_TEXT)
async def manager_back(message: Message) -> None:
    if not is_manager_or_admin(message.from_user.id):
        return
    await show_manager_menu(message)


@router.message()
async def manager_fallback(message: Message) -> None:
    if not is_manager_or_admin(message.from_user.id):
        return
    await message.answer(
        "Пожалуйста, выберите пункт меню.",
        reply_markup=_manager_main_menu_for(message.from_user.id),
    )
